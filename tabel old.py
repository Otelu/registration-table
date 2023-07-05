import sys
import os.path
import tkinter as tk
from tkinter import *
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import numbers, Alignment
import qrcode




# file creation

# Verifica daca fisierul exista #
file_exists = os.path.exists("Inscriere Sportivi.xlsx")


if file_exists is True:    # Daca exista ii da load si scrie in el
    wb = load_workbook("Inscriere Sportivi.xlsx")

else:     # Daca nu il creeaza
    wb = Workbook()

ws1 = wb.active
ws1.title = "Baieti"
ws2 = wb.create_sheet("Fete")


# Window
root = tk.Tk()
root.title('Aranjamente liste')

# Grid
canvas = tk.Canvas(root, width = 600, height= 300)
canvas.grid(columnspan = 3, rowspan = 20)


# Instrucions
instruction = tk.Label(root, text = "Scrieti CNP + NUME si PRENUME / CLUB ", font = "Raleway")
instruction.grid(columnspan = 3, column = 0, row = 0)


# CheckBox

## Variabile
var1 = IntVar()
var2 = IntVar()
var3 = IntVar()
var4 = IntVar()
var5 = IntVar()
var6 = IntVar()
var7 = IntVar()
var8 = IntVar()
var9 = IntVar()
var10 = IntVar()
var11 = IntVar()
var12 = IntVar()
var13 = IntVar()
var14 = IntVar()

# Definesc CheckBox
kihon_c = Checkbutton( root, text = "Kihon Dousa", variable = var1, onvalue = 1, offvalue = 0, command = lambda:comanda())
kodachi_c = Checkbutton( root, text = "Kodachi", variable = var2, onvalue = 1, offvalue = 0, command = lambda:comanda() )
tatekodachi_c = Checkbutton( root, text = "Tate Kodachi", variable = var3, onvalue = 1, offvalue = 0, command = lambda:comanda())
nito_c = Checkbutton( root, text = "Nito", variable = var4, onvalue = 1, offvalue = 0, command = lambda:comanda())
chokenFree_c = Checkbutton( root, text = "Choken Free", variable = var5, onvalue = 1, offvalue = 0, command = lambda:comanda())
chokenMorote_c = Checkbutton( root, text = "Choken Morote", variable = var6, onvalue = 1, offvalue = 0, command = lambda:comanda())
tanto_c = Checkbutton( root, text = "Tanto", variable = var7, onvalue = 1, offvalue = 0, command = lambda:comanda())
tateChoken_c = Checkbutton( root, text = "Tate Choken", variable = var8, onvalue = 1, offvalue = 0, command = lambda:comanda())
yari_c = Checkbutton( root, text = "Yari", variable = var9, onvalue = 1, offvalue = 0, command = lambda:comanda())
bo_c = Checkbutton( root, text = "Bo", variable = var10, onvalue = 1, offvalue = 0, command = lambda:comanda())
echipeLupta_c = Checkbutton( root, text = "Echipe Lupta", variable = var11, onvalue = 1, offvalue = 0, command = lambda:comanda())
echipeKhion_c = Checkbutton( root, text = "Echipe Kihon Dousa", variable = var12, onvalue = 1, offvalue = 0, command = lambda:comanda())
aviz_medical_c = Checkbutton(root, text="Aviz Medical", variable = var13, onvalue = 1, offvalue = 0, command = lambda:comanda())
stagiu = Checkbutton(root, text="Stagiu", variable = var14, onvalue = 1, offvalue = 0, command = lambda:comanda())



# Le aranjez
kihon_c.grid( columnspan = 1, column = 0, row = 3, sticky=W)
kodachi_c.grid( columnspan = 1, column = 0, row = 4, sticky=W)
tatekodachi_c.grid(columnspan = 1, column = 1, row = 3, sticky=W)
chokenMorote_c.grid(columnspan = 1, column = 1, row = 4, sticky=W)
chokenFree_c.grid(columnspan = 1, column = 2, row = 3, sticky=W)
nito_c.grid(columnspan = 1, column = 2, row = 4, sticky=W)
tanto_c.grid( columnspan = 1, column = 0, row = 5, sticky=W)
tateChoken_c.grid(columnspan = 1, column = 0, row = 6, sticky=W)
yari_c.grid( columnspan = 1, column = 1, row = 5, sticky=W)
bo_c.grid( columnspan = 1, column = 1, row = 6, sticky=W)
echipeLupta_c.grid( columnspan = 1, column = 2, row = 5, sticky=W)
echipeKhion_c.grid( columnspan = 1, column = 2, row = 6, sticky=W)
aviz_medical_c.grid(columnspan = 1, column = 1, row = 7)
stagiu.grid(columnspan = 1, column = 1, row = 8)


def comanda():
    print(var1.get())
    print(var2.get())


# ADD Button
add_button = tk.Button(root,command = lambda:add(ws1,ws2), text = "Adauga", font = "Raleway")
add_button.grid(columnspan=3,  column= 0, row = 15 )

# Quit Button
quit_button = tk.Button(root, command = lambda: sys.exit(), text = "Iesire", font = "Raleway")
quit_button.grid(columnspan = 3, column = 0, row = 20)



#### AICI BAGI TOATE FUNCTIILE ####
def add(ws1,ws2):


    nume = text.get(1.14, "2.0-1c")
    cnp = text.get(1.0, 1.13)
    global club
    club = text.get(2.0, 'end-1c')
    c_cnp = cnp



    # Check if the checkboxes are checked

    proba1 = "x"
    proba2 = "x"
    proba3 = "x"
    proba4 = "x"
    proba5 = "x"
    proba6 = "x"
    proba7 = "x"
    proba8 = "x"
    proba9 = "x"
    proba10 = "x"
    proba11 = "x"
    proba12 = "x"
    aviz = "Nu are aviz medical"


    if var1.get() == 1:
        proba1 = "Kihon Dousa"
        print(proba1)
    if var2.get() == 1:
        proba2 = "Kodachi"
        print(proba2)
    if var3.get() == 1:
        proba3 = "Tate Kodachi"
        print(proba3)
    if var4.get() == 1:
        proba4 = "Nito"
        print(proba4)
    if var5.get() == 1:
        proba5 = "Choken Free"
        print(proba5)
    if var6.get() == 1:
        proba6 = "Choken Morote"
        print(proba6)
    if var7.get() == 1:
        proba7 = "Tanto"
        print(proba7)
    if var8.get() == 1:
        proba8 = "Tate Choken"
        print(proba8)
    if var9.get() == 1:
        proba9 = "Yari"
        print(proba9)
    if var10.get() == 1:
        proba10 = "Bo"
        print(proba10)
    if var11.get() == 1:
        proba11 = "Echipe Lupta"
        print(proba11)
    if var12.get() == 1:
        proba12 = "Echipe Kihon Dousa"
        print(proba12)
    if var13.get() == 1:
        aviz = "Are aviz medical"
        print(aviz)


    varsta = int(c_cnp)/10000000000%100
    x = 1999 + varsta
    ani = 2022 - x

    an_nastere = int(c_cnp)/10000000000%100 + 2000
    luna_nastere = int(c_cnp)/100000000%100
    zi_nastere = int(c_cnp)/1000000%100

    # Data de nastere comparata cu data de azi
    data_azi = datetime.date.today()
    current_year = datetime.date.today().year
    data_sportiv = datetime.date(current_year, int(luna_nastere), int(zi_nastere))

    # Compara data de azi cu data de nastere a sportivului
    if data_sportiv > data_azi:
        ani = ani - 1

    ##CONTACT LEGITIMATIE QR
    contact = f'Nume: {nume}\nVarsta: {str(int(ani))}\nClub: {club}'

    ## AICI SCRIE IN FISIER ##
    if cnp.startswith('5') is True:
        print("merge")

        img = qrcode.make(contact)  ##### IMPORTANT
        img.save(nume + ".png")

        if var14.get() == 1:
            ws1.append([cnp, nume, club, str(int(ani)) + ' ani',str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), 'Stagiu'])
        else:
            ws1.append([cnp, nume, club, str(int(ani)) + ' ani', str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),aviz , str(proba1), str(proba2), str(proba3), str(proba4),
                str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11) ,str(proba12), '\n'])


    elif cnp.startswith('6') == True:
        print("merge fete")

        img = qrcode.make(contact)
        img.save(nume + ".png")

        if var14.get() == 1:
            ws1.append([cnp, nume, club, str(int(ani)) + ' ani',str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), 'Stagiu'])

        else:
            ws2.append([cnp, nume, club, str(int(ani)) + ' ani', str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),aviz ,str(proba1), str(proba2), str(proba3), str(proba4),
                        str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11) ,str(proba12), '\n'])


    ws1.auto_filter.add_filter_column(3,['ani'])
    ws1.auto_filter.add_sort_condition("D1:D150")

    wb.save("Inscriere Sportivi.xlsx")

    # Se sterge dupa adaugare
    text.delete(1.0, END)
    kihon_c.deselect()
    kodachi_c.deselect()
    nito_c.deselect()
    chokenFree_c.deselect()
    chokenMorote_c.deselect()
    tanto_c.deselect()
    tatekodachi_c.deselect()
    tateChoken_c.deselect()
    bo_c.deselect()
    yari_c.deselect()
    echipeKhion_c.deselect()
    echipeLupta_c.deselect()




# Text Box
text = Text(root, width = 40, height = 3, )
text.insert(INSERT, "")
text.grid(columnspan=3, column=0, row=1)



# Stilizare, arata mai bine
canvas = tk.Canvas(root, width = 600, height= 250)
canvas.grid(columnspan = 3)


## Mainloop of the window

root.mainloop()
