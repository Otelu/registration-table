import sys
import os.path
import tkinter as tk
from tkinter import *
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

# file creation

# Verifica daca fisierul exista #
file_exists = os.path.exists("Inscriere Sportivi.xlsx")

if file_exists is True:  # Daca exista ii da load si scrie in el
    wb = load_workbook("Inscriere Sportivi.xlsx")

else:  # Daca nu il creeaza
    wb = Workbook()


# Window
root = tk.Tk()
root.title('Aranjamente liste')

# Grid
canvas = tk.Canvas(root, width=600, height=300)
canvas.grid(columnspan=3, rowspan=20)

# Instrucions
instruction = tk.Label(root, text="Scrieti CNP + NUME si PRENUME / CLUB ", font="Raleway")
instruction.grid(columnspan=3, column=0, row=0)

# CheckBox

## Variabile
var1 = IntVar()
var2 = IntVar()
var3 = IntVar()
var4 = IntVar()
var5 = IntVar()
var6 = IntVar()
var7 = IntVar()
var8 = IntVar()
var9 = IntVar()
var10 = IntVar()
var_echipaLupta = IntVar()
var_echipaKihon = IntVar()
var13 = IntVar()
var_stagiu = IntVar()

# Definesc CheckBox
kihon_c = Checkbutton(root, text="Kihon Dousa", variable=var1, onvalue=1, offvalue=0, command=lambda: comanda())
kodachi_c = Checkbutton(root, text="Kodachi", variable=var2, onvalue=1, offvalue=0, command=lambda: comanda())
tatekodachi_c = Checkbutton(root, text="Tate Kodachi", variable=var3, onvalue=1, offvalue=0, command=lambda: comanda())
nito_c = Checkbutton(root, text="Nito", variable=var4, onvalue=1, offvalue=0, command=lambda: comanda())
chokenFree_c = Checkbutton(root, text="Choken Free", variable=var5, onvalue=1, offvalue=0, command=lambda: comanda())
chokenMorote_c = Checkbutton(root, text="Choken Morote", variable=var6, onvalue=1, offvalue=0,
                             command=lambda: comanda())
tanto_c = Checkbutton(root, text="Tanto", variable=var7, onvalue=1, offvalue=0, command=lambda: comanda())
tateChoken_c = Checkbutton(root, text="Tate Choken", variable=var8, onvalue=1, offvalue=0, command=lambda: comanda())
yari_c = Checkbutton(root, text="Yari", variable=var9, onvalue=1, offvalue=0, command=lambda: comanda())
bo_c = Checkbutton(root, text="Bo", variable=var10, onvalue=1, offvalue=0, command=lambda: comanda())
echipeLupta_c = Checkbutton(root, text="Echipe Lupta", variable=var_echipaLupta, onvalue=1, offvalue=0, command=lambda: comanda())
echipeKhion_c = Checkbutton(root, text="Echipe Kihon Dousa", variable=var_echipaKihon, onvalue=1, offvalue=0,
                            command=lambda: comanda())
aviz_medical_c = Checkbutton(root, text="Aviz Medical", variable=var13, onvalue=1, offvalue=0,
                             command=lambda: comanda())
stagiu = Checkbutton(root, text="Stagiu", variable=var_stagiu, onvalue=1, offvalue=0, command=lambda: comanda())

# Le aranjez
kihon_c.grid(columnspan=1, column=0, row=3, sticky=W)
kodachi_c.grid(columnspan=1, column=0, row=4, sticky=W)
tatekodachi_c.grid(columnspan=1, column=1, row=3, sticky=W)
chokenMorote_c.grid(columnspan=1, column=1, row=4, sticky=W)
chokenFree_c.grid(columnspan=1, column=2, row=3, sticky=W)
nito_c.grid(columnspan=1, column=2, row=4, sticky=W)
tanto_c.grid(columnspan=1, column=0, row=5, sticky=W)
tateChoken_c.grid(columnspan=1, column=0, row=6, sticky=W)
yari_c.grid(columnspan=1, column=1, row=5, sticky=W)
bo_c.grid(columnspan=1, column=1, row=6, sticky=W)
echipeLupta_c.grid(columnspan=1, column=2, row=5, sticky=W)
echipeKhion_c.grid(columnspan=1, column=2, row=6, sticky=W)
aviz_medical_c.grid(columnspan=1, column=1, row=7)
stagiu.grid(columnspan=1, column=1, row=8)


def comanda():
    print(var1.get())
    print(var2.get())


# ADD Button
add_button = tk.Button(root, command=lambda: add(), text="Adauga", font="Raleway")
add_button.grid(columnspan=3, column=0, row=15)

# Quit Button
quit_button = tk.Button(root, command=lambda: sys.exit(), text="Iesire", font="Raleway")
quit_button.grid(columnspan=3, column=0, row=20)

def remove_worksheet(sheet_name):
    # Check if the worksheet exists
    if sheet_name in wb.sheetnames:
        # Remove the worksheet
        wb.remove(wb[sheet_name])
        print(f"Worksheet '{sheet_name}' has been removed.")
    else:
        print(f"Worksheet '{sheet_name}' does not exist.")



#### AICI BAGI TOATE FUNCTIILE ####
def add():
    nume = text.get(1.14, "2.0-1c")
    cnp = text.get(1.0, 1.13)
    global club, wsf1, wsf2, wsf3, wsf4, wsf5, wsf6
    club = text.get(2.0, 'end-1c')
    c_cnp = cnp

    # Check if the checkboxes are checked

    proba1 = "x"
    proba2 = "x"
    proba3 = "x"
    proba4 = "x"
    proba5 = "x"
    proba6 = "x"
    proba7 = "x"
    proba8 = "x"
    proba9 = "x"
    proba10 = "x"
    proba11 = "x"
    proba12 = "x"
    aviz = "Nu are aviz medical"

    if var1.get() == 1:
        proba1 = "Kihon Dousa"
        print(proba1)
    if var2.get() == 1:
        proba2 = "Kodachi"
        print(proba2)
    if var3.get() == 1:
        proba3 = "Tate Kodachi"
        print(proba3)
    if var4.get() == 1:
        proba4 = "Nito"
        print(proba4)
    if var5.get() == 1:
        proba5 = "Choken Free"
        print(proba5)
    if var6.get() == 1:
        proba6 = "Choken Morote"
        print(proba6)
    if var7.get() == 1:
        proba7 = "Tanto"
        print(proba7)
    if var8.get() == 1:
        proba8 = "Tate Choken"
        print(proba8)
    if var9.get() == 1:
        proba9 = "Yari"
        print(proba9)
    if var10.get() == 1:
        proba10 = "Bo"
        print(proba10)
    if var_echipaLupta.get() == 1:
        proba11 = "Echipe Lupta"
        print(proba11)
    if var_echipaKihon.get() == 1:
        proba12 = "Echipe Kihon Dousa"
        print(proba12)
    if var13.get() == 1:
        aviz = "Are aviz medical"
        print(aviz)

    varsta = int(c_cnp) / 10000000000 % 100
    x = 1999 + varsta
    ani = datetime.date.today().year - x

    an_nastere = int(c_cnp) / 10000000000 % 100 + 2000
    luna_nastere = int(c_cnp) / 100000000 % 100
    zi_nastere = int(c_cnp) / 1000000 % 100

    # Data de nastere comparata cu data de azi
    data_azi = datetime.date.today()
    current_year = datetime.date.today().year
    data_sportiv = datetime.date(current_year, int(luna_nastere), int(zi_nastere))

    # Compara data de azi cu data de nastere a sportivului
    if data_sportiv > data_azi:
        ani = ani - 1

    sheet_names = wb.sheetnames

    ## AICI SCRIE IN FISIER ##

    ## STAGIU ##
    if var_stagiu.get() == 1:
        if "Stagiu" not in sheet_names:
            ws_stagiu = wb.create_sheet(title="Stagiu")
        else:
            ws_stagiu = wb["Stagiu"]

        ws_stagiu.append([cnp, nume, club, str(int(ani)) + ' ani',
                      str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz, '\n'])

        ## BAIETI ##
    elif cnp.startswith('5') is True:
        print("merge")
        if (ani < 21) and (ani > 18):
            if "Baieti 19-21" not in sheet_names:
                ws1 = wb.create_sheet(title="Baieti 19-21")
            else:
                ws1 = wb["Baieti 19-21"]

            ## Echipe ##
            if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                if "Echipe" not in sheet_names:
                    ws_echipe = wb.create_sheet(title="Echipe")
                else:
                    ws_echipe = wb["Echipe"]
                ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),
                            str(proba12), '\n' ])
                ws1.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), '\n'])
            ## CONCURS ##
            else:
                ws1.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                            str(proba12), '\n'])

        elif (ani < 18) and (ani > 16):
            if "Baieti 16-18" not in sheet_names:
                ws2 = wb.create_sheet(title="Baieti 16-18")
            else:
                ws2 = wb["Baieti 16-18"]

            ## ECHIPE ##
            if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                if "Echipe" not in sheet_names:
                    ws_echipe = wb.create_sheet(title="Echipe")
                else:
                    ws_echipe = wb["Echipe"]
                ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),
                            str(proba12), '\n' ])
                ws2.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), '\n'])

            else:
                ws2.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                            str(proba12), '\n'])

        elif (ani < 16) and (ani > 14):
            if "Baieti 14-16" not in sheet_names:
                ws3 = wb.create_sheet(title="Baieti 14-16")
            else:
                ws3 = wb["Baieti 14-16"]

            ## ECHIPE ##
            if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                if "Echipe" not in sheet_names:
                    ws_echipe = wb.create_sheet(title="Echipe")
                else:
                    ws_echipe = wb["Echipe"]
                ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),
                            str(proba12), '\n' ])
                ws3.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), '\n'])
            else:
                ws3.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                            str(proba12), '\n'])

        elif (ani < 14) and (ani > 12):
            if "Baieti 12-14" not in sheet_names:
                ws4 = wb.create_sheet(title="Baieti 12-14")
            else:
                ws4 = wb["Baieti 12-14"]

            ## ECHIPE ##
            if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                if "Echipe" not in sheet_names:
                    ws_echipe = wb.create_sheet(title="Echipe")
                else:
                    ws_echipe = wb["Echipe"]
                ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),
                            str(proba12), '\n' ])
                ws4.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10),  '\n'])
            else:
                ws4.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                            str(proba12), '\n'])

        elif (ani < 12) and (ani > 10):
            if "Baieti 10-12" not in sheet_names:
                ws5 = wb.create_sheet(title="Baieti 10-12")
            else:
                ws5 = wb["Baieti 10-12"]

            ## ECHIPE ##
            if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                if "Echipe" not in sheet_names:
                    ws_echipe = wb.create_sheet(title="Echipe")
                else:
                    ws_echipe = wb["Echipe"]
                ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),
                            str(proba12), '\n' ])
                ws5.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), '\n'])
            else:
                ws5.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                            str(proba12), '\n'])

        elif (ani < 10) and (ani > 7):
            if "Baieti 7-10" not in sheet_names:
                ws6 = wb.create_sheet(title="Baieti 7-10")
            else:
                ws6 = wb["Baieti 7-10"]

            ## ECHIPE ##
            if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                if "Echipe" not in sheet_names:
                    ws_echipe = wb.create_sheet(title="Echipe")
                else:
                    ws_echipe = wb["Echipe"]
                ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),
                            str(proba12), '\n' ])
                ws6.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10),'\n'])
            else:
                ws6.append([cnp, nume, club, str(int(ani)) + ' ani',
                            str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                            str(proba1), str(proba2), str(proba3), str(proba4),
                            str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                            str(proba12), '\n'])



    ## FETE ##
    elif cnp.startswith('6'):
        print("merge fete")
        if (ani < 21) and (ani > 18):
            if "Fete 19-21" not in sheet_names:
                wsf1 = wb.create_sheet(title="Fete 19-21")
            else:
                wsf1 = wb["Fete 19-21"]

           ## ECHIPE ##
            if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                if "Echipe" not in sheet_names:
                    ws_echipe = wb.create_sheet(title="Echipe")
                else:
                    ws_echipe = wb["Echipe"]
                ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),
                            str(proba12), '\n' ])
                wsf1.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                             str(proba1), str(proba2), str(proba3), str(proba4),
                             str(proba5), str(proba6), str(proba7), str(proba8), str(proba9),'\n'])

            else:
                wsf1.append([cnp, nume, club, str(int(ani)) + ' ani',
                         str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                         str(proba1), str(proba2), str(proba3), str(proba4),
                         str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                         str(proba12), '\n'])

        elif (ani < 18) and (ani > 16):
            if "Fete 16-18" not in sheet_names:
                wsf2 = wb.create_sheet(title="Fete 16-18")
            else:
                wsf2 = wb["Fete 16-18"]

                ## ECHIPE ##
                if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                    if "Echipe" not in sheet_names:
                        ws_echipe = wb.create_sheet(title="Echipe")
                    else:
                        ws_echipe = wb["Echipe"]
                    ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                                      str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),str(proba12), '\n'])
                    wsf2.append([cnp, nume, club, str(int(ani)) + ' ani',
                                 str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,str(proba1), str(proba2), str(proba3), str(proba4),
                                 str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10),'\n'])

                else:
                    wsf2.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                             str(proba1), str(proba2), str(proba3), str(proba4),
                             str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                             str(proba12), '\n'])

        elif (ani < 16) and (ani > 14):
            if "Fete 14-16" not in sheet_names:
                wsf3 = wb.create_sheet(title="Fete 14-16")
            else:
                wsf3 = wb["Fete 14-16"]

            ## ECHIPE ##
            if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                if "Echipe" not in sheet_names:
                    ws_echipe = wb.create_sheet(title="Echipe")
                else:
                    ws_echipe = wb["Echipe"]
                ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),
                            str(proba12), '\n' ])
                wsf3.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                             str(proba1), str(proba2), str(proba3), str(proba4),
                             str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), '\n'])
            else:
                wsf3.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                             str(proba1), str(proba2), str(proba3), str(proba4),
                             str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                             str(proba12), '\n'])

        elif (ani < 14) and (ani > 12):
            if "Fete 12-14" not in sheet_names:
                wsf4 = wb.create_sheet(title="Fete 12-14")
            else:
                wsf4 = wb["Fete 12-14"]

            ## ECHIPE ##
            if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                if "Echipe" not in sheet_names:
                    ws_echipe = wb.create_sheet(title="Echipe")
                else:
                    ws_echipe = wb["Echipe"]
                ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),
                            str(proba12), '\n' ])
                wsf4.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                             str(proba1), str(proba2), str(proba3), str(proba4),
                             str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10),'\n'])
            else:
                wsf4.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                             str(proba1), str(proba2), str(proba3), str(proba4),
                             str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                             str(proba12), '\n'])

        elif (ani < 12) and (ani > 10):
            if "Fete 10-12" not in sheet_names:
                wsf5 = wb.create_sheet(title="Fete 10-12")
            else:
                wsf5 = wb["Fete 10-12"]

            ## ECHIPE ##
            if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                if "Echipe" not in sheet_names:
                    ws_echipe = wb.create_sheet(title="Echipe")
                else:
                    ws_echipe = wb["Echipe"]
                ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),
                            str(proba12), '\n' ])
                wsf5.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                             str(proba1), str(proba2), str(proba3), str(proba4),
                             str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), '\n'])

            else:
                wsf5.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                             str(proba1), str(proba2), str(proba3), str(proba4),
                             str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                             str(proba12), '\n'])

        elif (ani < 10) and (ani > 7):
            if "Fete 7-10" not in sheet_names:
                wsf6 = wb.create_sheet(title="Fete 7-10")
            else:
                wsf6 = wb["Fete 7-10"]

            ## ECHIPE ##
            if var_echipaLupta.get() == 1 or var_echipaKihon.get() == 1:
                if "Echipe" not in sheet_names:
                    ws_echipe = wb.create_sheet(title="Echipe")
                else:
                    ws_echipe = wb["Echipe"]
                ws_echipe.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)),str(proba11),
                            str(proba12), '\n' ])
                wsf6.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                             str(proba1), str(proba2), str(proba3), str(proba4),
                             str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10),'\n'])
            else:
                wsf6.append([cnp, nume, club, str(int(ani)) + ' ani',
                             str(int(zi_nastere)) + '.' + str(int(luna_nastere)) + '.' + str(int(an_nastere)), aviz,
                             str(proba1), str(proba2), str(proba3), str(proba4),
                             str(proba5), str(proba6), str(proba7), str(proba8), str(proba9), str(proba10), str(proba11),
                             str(proba12), '\n'])




    remove_worksheet("Sheet")

    wb.save("Inscriere Sportivi.xlsx")
    wb.close()

    # Se sterge dupa adaugare
    text.delete(1.0, END)
    kihon_c.deselect()
    kodachi_c.deselect()
    nito_c.deselect()
    chokenFree_c.deselect()
    chokenMorote_c.deselect()
    tanto_c.deselect()
    tatekodachi_c.deselect()
    tateChoken_c.deselect()
    bo_c.deselect()
    yari_c.deselect()
    echipeKhion_c.deselect()
    echipeLupta_c.deselect()
    aviz_medical_c.deselect()
    stagiu.deselect()


# Text Box
text = Text(root, width=40, height=3, )
text.insert(INSERT, "")
text.grid(columnspan=3, column=0, row=1)

# Stilizare, arata mai bine
canvas = tk.Canvas(root, width=600, height=250)
canvas.grid(columnspan=3)

## Mainloop of the window

root.mainloop()
